"""XLSX/CSV генерация. Используется и из бота, и из API (через .delay()).

`build_xlsx_bytes` — синхронная, чтобы вызывать из любого контекста.
"""

from __future__ import annotations

import csv
import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from packages.db.models import Expense, Project
from packages.domain.categories import label_for
from packages.domain.reports import summarize_expenses


def build_xlsx_bytes(project: Project, expenses: Iterable[Expense], *, locale: str = "ru") -> bytes:
    wb = Workbook()
    expenses_list = list(expenses)

    # Sheet 1: расходы
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Расходы" if locale == "ru" else "Expenses"
    headers = (
        ["Дата", "Категория", "Сумма", "Валюта", "Описание", "Источник"]
        if locale == "ru"
        else ["Date", "Category", "Amount", "Currency", "Description", "Source"]
    )
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EEEEEE")

    for e in expenses_list:
        ws.append([
            e.paid_at.isoformat(),
            label_for(e.category, locale),  # type: ignore[arg-type]
            e.amount_minor / 100.0,
            e.currency,
            e.description or "",
            e.source,
        ])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["F"].width = 12

    # Sheet 2: сводка по категориям
    summary = summarize_expenses(expenses_list)
    ws2 = wb.create_sheet("Категории" if locale == "ru" else "Categories")
    cat_headers = (
        ["Категория", "Сумма", "Число чеков"]
        if locale == "ru"
        else ["Category", "Total", "Count"]
    )
    ws2.append(cat_headers)
    for c in ws2[1]:
        c.font = Font(bold=True)
    for row in summary.by_category:
        ws2.append([
            label_for(row.category, locale),  # type: ignore[arg-type]
            row.total_minor / 100.0,
            row.count,
        ])
    ws2.append([])
    ws2.append([
        "Итого" if locale == "ru" else "Total",
        summary.total_minor / 100.0,
        summary.count,
    ])
    for c in ws2[ws2.max_row]:
        c.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    # Sheet 3: header сводка
    ws3 = wb.create_sheet("Сводка" if locale == "ru" else "Summary")
    ws3.append([("Проект" if locale == "ru" else "Project"), project.name])
    ws3.append([("Валюта" if locale == "ru" else "Currency"), project.currency])
    ws3.append([("Итого" if locale == "ru" else "Total"), summary.total_minor / 100.0])
    ws3.append([("Число чеков" if locale == "ru" else "Receipt count"), summary.count])
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 40
    for row in ws3.iter_rows():
        row[0].font = Font(bold=True)
        row[0].alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv_bytes(project: Project, expenses: Iterable[Expense], *, locale: str = "ru") -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    headers = (
        ["date", "category", "amount", "currency", "description", "source"]
    )
    w.writerow(headers)
    for e in expenses:
        w.writerow([
            e.paid_at.isoformat(),
            e.category,
            f"{e.amount_minor / 100.0:.2f}",
            e.currency,
            e.description or "",
            e.source,
        ])
    return buf.getvalue().encode("utf-8")
